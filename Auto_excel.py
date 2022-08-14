from openpyxl import Workbook
import openpyxl

wb = Workbook()
ws = wb.create_sheet("A sheet", 0)
ws2 = wb.create_sheet("Sheet nr 2")

# for sheet in wb:
#     print(sheet.title)

wb.remove(wb["Sheet"])
# print("------------")
# for sheet in wb:
#     print(sheet.title)

print(ws2.title)
ws2.title = "New title"

for sheet in wb:
    print(sheet.title)
